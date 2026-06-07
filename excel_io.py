"""
excel_io.py
-----------
Multi-sheet Excel okuma/yazma. Atomik yazma garantisi:
1. Lock file kontrol
2. Once .tmp uzantili dosyaya yaz
3. Basariliysa rename (os.replace - atomik)

Bir GroupExcel instance bir Excel dosyasini (bir coin grubu) yonetir.
Sheet adi format: BTC_verileri, ETH_verileri vs.

Kullanicinin orijinal sutun semasi tam destekleniyor:
tarih, seans, btc_trendi, korku_endeksi, tahta_al_sat_orani, cvd_miktari,
oi_degisimi, funding_oran, duvar_mesafesi, iptal_edilen_duvar_orani,
btc_korelasyonu, poc_uzakligi, likidasyon_miktari, vwap_sapmasi, rsi_degeri,
tahmin_yonu, giris_fiyati, hedef_fiyat, stop_fiyati, islem_durumu,
kirmizi_set_etiketi

Buna ek olarak ~25 ek sutun (skor, confidence, ATR, multi-TF degerleri vb.).
"""
from __future__ import annotations

import os
import time
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src import config
from src.utils import setup_logger, to_sheet_name

logger = setup_logger("excel_io")


# ============= ANA SUTUN SEMASI =============
# Kullanicinin orijinal sartnamesindeki 21 sutun + ek detaylar.
# Sira korunmali - Excel'i actiginda once ana sutunlar gozukur.

COLUMN_SCHEMA: list[str] = [
    # --- KULLANICININ 21 SUTUN (ORIJINAL SARTNAME) ---
    "tarih",
    "seans",
    "btc_trendi",
    "korku_endeksi",
    "tahta_al_sat_orani",
    "cvd_miktari",
    "oi_degisimi",
    "funding_oran",
    "duvar_mesafesi",
    "iptal_edilen_duvar_orani",
    "btc_korelasyonu",
    "poc_uzakligi",
    "likidasyon_miktari",
    "vwap_sapmasi",
    "rsi_degeri",
    "tahmin_yonu",
    "giris_fiyati",
    "hedef_fiyat",
    "stop_fiyati",
    "islem_durumu",
    "kirmizi_set_etiketi",
    # --- EK META ---
    "skor",
    "confidence",
    "atr_kullanildi",
    "pozisyon_buyukluk_usd",
    "cikis_fiyati",
    "kapanis_tarihi",
    "pnl_pct",
    # --- EK FEATURE DETAYLARI ---
    "coin_trend_15m",
    "coin_trend_1h",
    "pct_change_15m",
    "pct_change_1h",
    "pct_change_24h",
    "rsi_5m",
    "rsi_1h",
    "macd_hist_15m",
    "bb_pct_b_15m",
    "bb_width_15m",
    "atr_pct_15m",
    "imb_0p5",
    "imb_2p0",
    "spread_bps",
    "bid_wall_size",
    "ask_wall_size",
    "bid_wall_dist_pct",
    "ask_wall_dist_pct",
    "largest_wall_side",
    "whale_count_15m",
    "whale_net_flow_15m",
    "top_trader_ls_pos",
    "global_ls_ratio",
    "taker_ratio_5m",
    "taker_ratio_15m",
    "vwap_daily",
    "poc_24h",
    "vah_24h",
    "val_24h",
    "btc_dominance",
    "btc_pct_1h",
    "fng_7d_avg",
    "realized_vol_24h",
    "open_interest_usd",
    "funding_24h_avg",
    "errors_count",
]


# ============= LOCK FILE =============
class FileLock:
    """Basit file lock - aynı dosyaya iki cron eş zamanlı yazmasin diye."""

    def __init__(self, lock_path: Path, timeout_sec: int = 30) -> None:
        self.lock_path = lock_path
        self.timeout_sec = timeout_sec

    def acquire(self) -> bool:
        """Lock alir. Eski (>5dk) lock varsa kaldirir."""
        if self.lock_path.exists():
            try:
                age = time.time() - self.lock_path.stat().st_mtime
                if age > 300:  # 5 dakika - stale lock
                    self.lock_path.unlink()
                    logger.warning(f"Eski lock kaldirildi: {self.lock_path}")
                else:
                    return False
            except OSError:
                return False
        try:
            self.lock_path.parent.mkdir(parents=True, exist_ok=True)
            with open(self.lock_path, "w") as f:
                f.write(str(os.getpid()))
            return True
        except OSError:
            return False

    def release(self) -> None:
        try:
            if self.lock_path.exists():
                self.lock_path.unlink()
        except OSError:
            pass

    def __enter__(self):
        if not self.acquire():
            raise RuntimeError(f"Lock alinamadi: {self.lock_path}")
        return self

    def __exit__(self, *args):
        self.release()


# ============= GROUP EXCEL MANAGER =============
class GroupExcel:
    """
    Bir coin grubunun Excel dosyasini yonetir.
    Acilis: dosya yoksa olusturulur, tum sheet'ler bos.
    Sheet manipulation: in-memory DataFrame dict.
    Yazim: write_all() ile atomik kayit.
    """

    def __init__(self, group: str) -> None:
        self.group = group
        self.path = config.excel_path_for_group(group)
        self.coins = config.COIN_GROUPS[group]
        self.lock_path = self.path.with_suffix(".lock")
        self.logger = logger
        # sheet -> DataFrame
        self.sheets: dict[str, pd.DataFrame] = {}

    def load(self) -> None:
        """Diskten oku. Yoksa bos DataFrame'ler hazirla."""
        if not self.path.exists():
            self.logger.info(f"Yeni Excel dosyasi olusturuluyor: {self.path.name}")
            for coin in self.coins:
                sheet = to_sheet_name(coin)
                self.sheets[sheet] = pd.DataFrame(columns=COLUMN_SCHEMA)
            return

        try:
            xl = pd.ExcelFile(self.path, engine="openpyxl")
            for coin in self.coins:
                sheet = to_sheet_name(coin)
                if sheet in xl.sheet_names:
                    df = pd.read_excel(xl, sheet_name=sheet)
                    # Eksik sutunlari ekle (yeni surumde kolon eklenmis olabilir)
                    for col in COLUMN_SCHEMA:
                        if col not in df.columns:
                            df[col] = pd.NA
                    # Sutun sirasini schema'ya gore yeniden duzenle (varsa ekstralari sona koy)
                    extra = [c for c in df.columns if c not in COLUMN_SCHEMA]
                    df = df[COLUMN_SCHEMA + extra]
                    self.sheets[sheet] = df
                else:
                    self.sheets[sheet] = pd.DataFrame(columns=COLUMN_SCHEMA)
        except Exception as e:
            self.logger.error(f"Excel okuma hatasi {self.path}: {e}")
            # Fallback: bos sheet'ler
            for coin in self.coins:
                sheet = to_sheet_name(coin)
                if sheet not in self.sheets:
                    self.sheets[sheet] = pd.DataFrame(columns=COLUMN_SCHEMA)

    def get_df(self, coin: str) -> pd.DataFrame:
        """Bir coinin DataFrame'ini doner (referans, in-place degisikliklere izinli)."""
        sheet = to_sheet_name(coin)
        if sheet not in self.sheets:
            self.sheets[sheet] = pd.DataFrame(columns=COLUMN_SCHEMA)
        return self.sheets[sheet]

    def set_df(self, coin: str, df: pd.DataFrame) -> None:
        """Bir coinin DataFrame'ini gunceller (pending kontrol sonrasi)."""
        sheet = to_sheet_name(coin)
        # Tum schema kolonlari mevcut mu kontrol
        for col in COLUMN_SCHEMA:
            if col not in df.columns:
                df[col] = pd.NA
        self.sheets[sheet] = df

    def append_row(self, coin: str, row: dict[str, Any]) -> None:
        """Bir coin sheet'ine yeni satir ekler. Eksik kolonlar NaN."""
        df = self.get_df(coin)
        # Sadece schema'daki kolonlari al, eksiklere NaN
        clean_row = {col: row.get(col, pd.NA) for col in df.columns}
        # Schema disinda extra kolon varsa ekle
        for k, v in row.items():
            if k not in clean_row:
                clean_row[k] = v
                df[k] = pd.NA
        df.loc[len(df)] = clean_row
        self.sheets[to_sheet_name(coin)] = df

    def write_atomic(self) -> bool:
        """
        Tum sheet'leri tek Excel'e atomik yazar.
        1. Lock al
        2. .tmp dosyaya yaz
        3. os.replace ile rename
        """
        with FileLock(self.lock_path, timeout_sec=30) as _lock:
            tmp_path = self.path.with_suffix(".xlsx.tmp")
            try:
                # ExcelWriter ile multi-sheet yaz
                with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                    for sheet_name, df in self.sheets.items():
                        if df is None or df.empty:
                            # Bos sheet'i de olustur ki coin sayfasi gozuksun
                            pd.DataFrame(columns=COLUMN_SCHEMA).to_excel(
                                writer, sheet_name=sheet_name, index=False
                            )
                        else:
                            df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Sutun basliklarini formatla
                self._format_headers(tmp_path)

                # Atomik rename
                os.replace(tmp_path, self.path)
                self.logger.info(
                    f"Excel kaydedildi: {self.path.name} "
                    f"({sum(len(d) for d in self.sheets.values())} satir)"
                )
                return True
            except Exception as e:
                self.logger.error(f"Excel yazma hatasi {self.path}: {e}")
                if tmp_path.exists():
                    try:
                        tmp_path.unlink()
                    except OSError:
                        pass
                return False

    @staticmethod
    def _format_headers(path: Path) -> None:
        """Excel header'larina renk + bold uygula."""
        try:
            wb = load_workbook(path)
            header_fill = PatternFill("solid", start_color="1F4E78")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            center = Alignment(horizontal="center", vertical="center")
            for ws in wb.worksheets:
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center
                # Sutun genisligi
                for col in ws.columns:
                    max_len = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            val_len = len(str(cell.value))
                            if val_len > max_len:
                                max_len = val_len
                        except Exception:
                            pass
                    ws.column_dimensions[column].width = min(max_len + 2, 22)
                # Donduran ilk satir
                ws.freeze_panes = "A2"
            wb.save(path)
        except Exception as e:
            logger.warning(f"Header format hata: {e}")
